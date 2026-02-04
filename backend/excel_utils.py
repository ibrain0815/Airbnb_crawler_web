"""
엑셀 유틸 — 수집 결과를 엑셀 bytes 로 생성 (파일 시스템 저장 없음).
컬럼: 번호, 제목, 가격, 주소, 평점/후기, 링크.
"""

import logging
from datetime import datetime
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

logger = logging.getLogger(__name__)

# 엑셀 컬럼 순서 및 한글 헤더 (1행)
EXCEL_COLUMNS = [
    ("no", "번호"),
    ("title", "숙소명"),
    ("price", "가격"),
    ("address", "상세설명"),
    ("rating", "평점/후기"),
    ("url", "링크"),
]


def _apply_formatting(ws: Any) -> None:
    """헤더·데이터 행 서식 적용."""
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    data_align = Alignment(vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = data_align
            cell.border = border

    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            try:
                v = str(cell.value) if cell.value else ""
                n = sum(2 if ord(c) > 127 else 1 for c in v)
                max_len = max(max_len, n)
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 50)

    ws.freeze_panes = "A2"


def save_listings_to_excel(listings: list[dict]) -> bytes:
    """
    수집된 숙소 목록을 엑셀 파일 bytes 로 생성 (디스크 저장 없음).
    파일명 규칙: airbnb_listings_{timestamp}.xlsx (반환값에는 미포함).
    """
    if not listings:
        logger.warning("저장할 데이터가 없습니다.")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "목록"
        for c, (_, label) in enumerate(EXCEL_COLUMNS, 1):
            ws.cell(row=1, column=c, value=label)
        _apply_formatting(ws)
        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "목록"

    for c, (key, label) in enumerate(EXCEL_COLUMNS, 1):
        ws.cell(row=1, column=c, value=label)

    for r, item in enumerate(listings, 2):
        for c, (key, _) in enumerate(EXCEL_COLUMNS, 1):
            val = item.get(key, "")
            if isinstance(val, (list, dict)):
                val = str(val)
            ws.cell(row=r, column=c, value=val)

    _apply_formatting(ws)
    buf = BytesIO()
    wb.save(buf)
    logger.info("엑셀 생성 완료: %d행", len(listings))
    return buf.getvalue()


def get_excel_filename() -> str:
    """다운로드용 파일명: airbnb_listings_{timestamp}.xlsx"""
    return f"airbnb_listings_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
